from tkinter import *
from datetime import date
from tkinter import messagebox
from tkinter import filedialog
from PIL import Image,ImageTk
import os 
from tkinter.ttk import Combobox
import openpyxl
from openpyxl import Workbook
import pathlib 

background="#06283D"
framebg="#808080"
framefg="#06283D"
root=Tk()
root.title("STUDENT REGISTRATION SYSETEM")
root.geometry("1250x700+210+100")
root.config(bg=background)
file=pathlib.Path('Student_data.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="Registration No."
    sheet['B1']="Name "
    sheet['C1']="Class"
    sheet['D1']="Date Of Birth"
    sheet['E1']="Gender"
    sheet['F1']="Date of Registration"
    sheet['G1']="Religion"
    sheet['H1']="Skills"
    sheet['I1']="Father's name"
    sheet['J1']="Mother's Name"
    sheet['K1']="Father's occupation"
    sheet['L1']="Mother's occupation"

    file.save('Student_data.xlsx')

# Function to open file dialog and display selected image
def showimage():
    global img
    global lbl
    filename = filedialog.askopenfilename(initialdir=os.getcwd(), title="Browse to your profile section:",filetypes=(("JPG File", "*.jpg"), ("All File", "*.*"), ("PNG File", "*.png")))
    img = Image.open(filename)
    resized_img = img.resize((190, 190))
    photo2 = ImageTk.PhotoImage(resized_img)
    lbl.config(image=photo2)
    lbl.image = photo2
def reset():
    Name.set('')
    dob.set('')
    religion.set('')
    Class.set('select class')
    fathername.set("")
    mothername.set("")
    fatheroccupation.set("")
    motheroccupation.set("")
    skill.set('')
    
    registration1()
    img = PhotoImage(file="avatar.png")
    lbl.image=img
    lbl.config(image=img)
    img=''
def save():
    rid=registration.get()
    name=Name.get()
    currentdate=Date.get()
    birth=dob.get()
    grade=Class.get()
    Religion=religion.get()
    excelin=skill.get()
    sex="0"
    fname=fathername.get()
    mname=mothername.get()
    fjob=fatheroccupation.get()
    mjob=motheroccupation.get()
    if mjob=='' or fjob=='' or Religion=='' or fname=='' or mname=='' or name=='':
        messagebox.showerror('error',"there is empty field")
    else:
        file=openpyxl.load_workbook('Student_data.xlsx')
        sheet=file.active
        sheet.cell(column=1,row=sheet.max_row+1,value=rid)
        sheet.cell(column=2,row=sheet.max_row,value=name)
        sheet.cell(column=3,row=sheet.max_row,value=grade)
        sheet.cell(column=4,row=sheet.max_row,value=birth)
        sheet.cell(column=5,row=sheet.max_row,value=sex)
        sheet.cell(column=6,row=sheet.max_row,value=currentdate)
        sheet.cell(column=7,row=sheet.max_row,value=Religion)
        sheet.cell(column=8,row=sheet.max_row,value=excelin)
        sheet.cell(column=9,row=sheet.max_row,value=fname)
        sheet.cell(column=10,row=sheet.max_row,value=fjob)
        sheet.cell(column=11,row=sheet.max_row,value=mname)
        sheet.cell(column=12,row=sheet.max_row,value=mjob)
        
        file.save(r'Student_data.xlsx')
        
        try:
            img.save("student images/"+str(rid)+'.jpg')
        except Exception as e:
            os.mkdir("student images") 
            img.save("student images/"+str(rid)+'.jpg')
        messagebox.showinfo("info","successfully registered")
        reset()
        registration1()
from tkinter import messagebox
import openpyxl
from PIL import Image, ImageTk

def searchx():
    text = search.get()
    reset()
    savebtn.config(state=DISABLED)
    file = openpyxl.load_workbook("Student_data.xlsx")
    try:
        sheet = file.active
        for row in sheet.iter_rows(min_row=2, max_col=1):
            if row[0].value == text:
                regno = row[0].row
                break
        else:
            messagebox.showerror("Error", "Not found")
            return
    except Exception as e:
        return
    try:
        x1 = sheet.cell(row=regno, column=1).value
        x2 = sheet.cell(row=regno, column=2).value
        x3 = sheet.cell(row=regno, column=3).value
        x4 = sheet.cell(row=regno, column=5).value
        x5 = sheet.cell(row=regno, column=4).value
        x6 = sheet.cell(row=regno, column=6).value
        x7 = sheet.cell(row=regno, column=7).valuen
        x8 = sheet.cell(row=regno, column=8).value
        x9 = sheet.cell(row=regno, column=9).value
        x10 = sheet.cell(row=regno, column=10).value
        x11 = sheet.cell(row=regno, column=11).value
        x12 = sheet.cell(row=regno, column=12).value
        registration.set(x1)
        Name.set(x2)
        religion.set(x7)
        Class.set(x3)
        fathername.set(x9)
        mothername.set(x11)
        fatheroccupation.set(x10)
        motheroccupation.set(x12)
        radio.set(x4)
        dob.set(x5)
        skill.set(x8)
        img = Image.open("student images/" + str(x1) + ".jpg")
        resized_img = img.resize((190, 190))
        photo2 = ImageTk.PhotoImage(resized_img)
        lbl.config(image=photo2)
        lbl.image = photo2
    except Exception as e:
        messagebox.showerror("Error", "Error while retrieving data")

def uppdate():
    savebtn.config(state="active")
    rid = registration.get()
    name = Name.get()
    currentdate = Date.get()
    birth = dob.get()
    grade = Class.get()
    Religion = religion.get()
    excelin = skill.get()
    sex = "male"
    fname = fathername.get()
    mname = mothername.get()
    fjob = fatheroccupation.get()
    mjob = motheroccupation.get()

    if mjob == '' or fjob == '' or Religion == '' or fname == '' or mname == '' or name == '':
        messagebox.showerror('Error', 'There is an empty field')
        return
    try:
        file = openpyxl.load_workbook('Student_data.xlsx')
        sheet = file.active
        for row in sheet.iter_rows(min_row=2, max_col=1):
            if row[0].value == rid:
                regno = row[0].row
                break
        else:
            messagebox.showerror("Error", "Not found")
            return
        sheet.cell(row=regno, column=2, value=name)
        sheet.cell(row=regno, column=3, value=grade)
        sheet.cell(row=regno, column=4, value=birth)
        sheet.cell(row=regno, column=5, value=sex)
        sheet.cell(row=regno, column=6, value=currentdate)
        sheet.cell(row=regno, column=7, value=Religion)
        sheet.cell(row=regno, column=8, value=excelin)
        sheet.cell(row=regno, column=9, value=fname)
        sheet.cell(row=regno, column=10, value=fjob)
        sheet.cell(row=regno, column=11, value=mname)
        sheet.cell(row=regno, column=12, value=mjob)
        file.save("Student_data.xlsx")
        try:
            img.save("student images/" + str(rid) + ".jpg")
        except Exception as e:
            pass
        messagebox.showinfo("Update", "Updated successfully")
        reset()
    except Exception as e:
        messagebox.showerror("Error", "Error while updating data")
#registration and details
def registration1():
    try:
        file = openpyxl.load_workbook("Student_data.xlsx")
        sheet = file.active
        row = sheet.max_row
        max_row_value = int(sheet.cell(row=row, column=1).value)
        # Check if max_row_value is None or not an integer
        if max_row_value is None or not isinstance(max_row_value, int):
            registration.set(1)
        else:
            registration.set(max_row_value + 1)
    except Exception as e:
        registration.set(1)
Label(root, text="Registration No.:", font="Arial 13 bold", bg=background).place(x=30, y=150)
registration = StringVar()
registration1()
registration_entry = Entry(root, textvariable=registration,state=DISABLED, width=15, font="Arial 10")
registration_entry.place(x=180, y=155)
# Call the registration1 function to set the initial value
#top Levels
Label(root,text="your mail",width=10,height=3,bg="#f0867c",anchor='e').pack(side=TOP,fill=X)
Label(root,text="STUDENT REGISTRATION SYSTEM",width=10,height=2,bg="#fff",fg="#f0867c",font="ariel 20 bold").pack(side=TOP,fill=X)
savebtn=Button(root,text="Save",width=19,height=2,font="ariel 12 bold", bg="lightgreen",command=save)
savebtn.place(x=1000,y=450)
search=StringVar()
Entry(root,textvariable=search,width=15,bd=2,font="ariel 20").place(x=875,y=70)
imageicon=PhotoImage(file='search.png')
srch=Button(root,text="search",image=imageicon,width=123,bg="#68ddfa", font="ariel 13 bold",command=searchx)
srch.place(x=1111,y=70)
imageicon1=PhotoImage(file='update.png')
update=Button(root,text="update",image=imageicon1,width=50,height=30,bg="#68ddfa", font="ariel 13 bold",command=uppdate)
update.place(x=110,y=70)
Label(root, text="Date:", font="Arial 13 bold", bg=background).place(x=370, y=150)
Date=StringVar()
today=date.today()
day1=today.strftime('%d/%m/%Y')
date_entry=Entry(root,textvariable=Date,width=15,state=DISABLED  ,font="ariel 10").place(x=432,y=155)
Date.set(day1)
#student detials
obj=LabelFrame(root,text="Student's details ", font=20,bd=2,bg=framebg,fg=framefg,width=900,height=250,relief=GROOVE)
obj.place(x=30,y=200)
Label(obj,text="Full Name : ",font="ariel  13",bg=framebg,fg=framefg).place(x=30,y=50)
Label(obj,text="Date Of Birth : ",font="ariel  13",bg=framebg,fg=framefg).place(x=30,y=100)
Label(obj,text="Gender : ",font="ariel  13",bg=framebg,fg=framefg).place(x=30,y=150)
Label(obj,text="Class : ",font="ariel  13",bg=framebg,fg=framefg).place(x=500,y=50)
Label(obj,text="Religion  : ",font="ariel  13",bg=framebg,fg=framefg).place(x=500,y=100)
Label(obj,text="Skill : ",font="ariel  13",bg=framebg,fg=framefg).place(x=500,y=150)
def selection():
    global gender
    value = radio.get()
    if value == "Male":
        gender = "Male"
    else:
        gender = "Female"
Name = StringVar()
name_entry = Entry(obj, textvariable=Name, width=20, font="ariel 10")
name_entry.place(x=160, y=50)
dob = StringVar()
dob_entry = Entry(obj, textvariable=dob, width=20, font="ariel 10")
dob_entry.place(x=160, y=100)
radio = IntVar()
r1 = Radiobutton(obj, text="Male", variable=radio, value='Male', command=selection)
r1.place(x=150, y=150)
r1 = Radiobutton(obj, text="Female", variable=radio, value='Female', command=selection)
r1.place(x=200, y=150)
religion = StringVar()
religion_entry = Entry(obj, textvariable=religion, width=20, font="ariel 10")
religion_entry.place(x=630, y=100)
Class = Combobox(obj, values=["computer engineering", "civil engineering", "mechanical engineering", "architecture", "electrical engineering", "electronic engineering", "Bsc.CSIT", "BIM", "BCA"])
Class.place(x=630, y=50)
Class.set("select class")
Class.config(state="readonly") 
dob = StringVar()
dob_entry = Entry(obj, textvariable=dob, width=20, font="ariel 10")
dob_entry.place(x=160, y=100)
skill = StringVar()
skill_entry = Entry(obj, textvariable=skill, width=20, font="ariel 10")
skill_entry.place(x=630, y=150)
#parents detials
obj2=LabelFrame(root,text="Parents's details ", font=20,bd=2,bg=framebg,fg=framefg,width=900,height=200,relief=GROOVE)
obj2.place(x=30,y=470)
Label(obj2,text="Father's Name : ",font="ariel 13", bg=framebg,fg=framefg).place(x=30,y=50)
Label(obj2,text="Father's occupation : ",font="ariel 13", bg=framebg,fg=framefg).place(x=30,y=100)
Label(obj2,text="Mother's Name : ",font="ariel 13", bg=framebg,fg=framefg).place(x=500,y=50)
Label(obj2,text="Mother's occupation : ",font="ariel 13", bg=framebg,fg=framefg).place(x=500,y=100)
fathername=StringVar()
fathername_entry=Entry(obj2,textvariable=fathername,width=20,font="ariel 10")
fathername_entry.place(x=200,y=50)
fatheroccupation=StringVar()
fatheroccupation_entry=Entry(obj2,textvariable=fatheroccupation,width=20,font="ariel 10")
fatheroccupation_entry.place(x=200,y=100)
mothername=StringVar()
mothername_entry=Entry(obj2,textvariable=mothername,width=20,font="ariel 10")
mothername_entry.place(x=670,y=50)
motheroccupation=StringVar()
motheroccupation_entry=Entry(obj2,textvariable=motheroccupation,width=20,font="ariel 10")
motheroccupation_entry.place(x=670,y=100)
# Frame for displaying image
f = Frame(root, bd=3, bg="black", width=200, height=200, relief=GROOVE)
f.place(x=1000, y=150)
# Placeholder image
img = PhotoImage(file="avatar.png")
lbl = Label(f, bg="black", image=img, width=200, height=200)
lbl.place(x=0, y=0)
Button(root,text="Upload",width=19,height=2,font="ariel 12 bold", bg="lightblue",command=showimage).place(x=1000,y=370)
Button(root,text="Reset",width=19,height=2,font="ariel 12 bold", bg="lightpink",command=reset).place(x=1000,y=530)
def Exit():
    root.destroy()
Button(root,text="Exit",width=19,height=2,font="ariel 12 bold", bg="grey",command=Exit).place(x=1000,y=610)
root.mainloop()